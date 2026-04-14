# Copyright (c) 2014 Czech National Corpus
#
# This program is free software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; version 2
# dated June, 1991.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA  02110-1301, USA.

"""
A plug-in allowing export of a concordance (in fact, any row/cell
like data can be used) to XLSX (Office Open XML) format.

Plug-in requires openpyxl library.
"""
from io import BytesIO
from typing import Any, Dict, List, Tuple

from action.argmapping.wordlist import WordlistSaveFormArgs
from action.model.concordance import ConcActionModel
from action.model.keywords import KeywordsActionModel
from action.model.pquery import ParadigmaticQueryActionModel
from action.model.wordlist import WordlistActionModel
from babel import Locale
from babel.numbers import format_decimal
from bgcalc.coll_calc import CalculateCollsResult
from bgcalc.keywords import CNCKeywordLine, KeywordsResult
from bgcalc.pquery.storage import PqueryDataLine
from conclib.errors import ConcordanceQueryParamsError
from kwiclib.common import KwicPageData
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from views.colls import SavecollArgs
from views.concordance import SaveConcArgs
from views.freqs import SavefreqArgs
from views.keywords import SaveKeywordsArgs
from views.pquery import SavePQueryArgs

from . import AbstractExport, ExportPluginException, lang_row_to_list


class XLSXExport(AbstractExport):

    def __init__(self, locale: Locale):
        super().__init__(locale)
        self._written_lines = 0
        self._wb = Workbook(write_only=True)
        self._sheet = self._wb.create_sheet()
        self._col_types = ()
        self._import_row = lambda x: x

    def _formatnumber(self, x):
        if x is None or x == '':
            return ''
        try:
            return format_decimal(x, locale=self._locale, group_separator=False)
        except Exception:
            return str(x)

    def content_type(self):
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def raw_content(self):
        output = BytesIO()
        self._wb.save(filename=output)
        return output.getvalue()

    def _writeheading(self, data):
        if len(data) > 1 and data[0] != '' and all(s == '' for s in data[1:]):
            self._sheet.append([data[0]])
            for _ in range(3):
                self._sheet.append([])
            # this kind of a hack in "write-only" mode
            self._sheet.merged_cells.ranges.append('A1:G4')
        else:
            self._sheet.append(data)
        self._sheet.append([])
        self._written_lines += 2

    def _write_ref_headings(self, data):
        cells = []
        for d in data:
            cell = WriteOnlyCell(self._sheet, d)
            cell.font = cell.font.copy(bold=True)
            cells.append(cell)
        self._sheet.append(cells)
        self._written_lines += 1

    def _set_col_types(self, *types):
        self._col_types = types

    def _import_value(self, v, i):
        format_map = {
            int: '0',
            int: '0',
            float: '0.00',
            str: 'General'
        }
        if i < len(self._col_types):
            out_type = self._col_types[i]
        else:
            out_type = str
        if out_type not in format_map:
            raise ExportPluginException('Unsupported cell type %s' % out_type)
        if out_type is not str and v is None or v == '':
            out_type = str
        return out_type(v), format_map[out_type]

    def _writerow(self, line_num, *lang_rows):
        row = []
        if line_num is not None:
            row.append(line_num)
        for lang_row in lang_rows:
            row += self._import_row(lang_row)
        self._sheet.append([self._get_cell(*self._import_value(d, i)) for i, d in enumerate(row)])
        self._written_lines += 1

    def _new_sheet(self, name):
        if self._written_lines > 1:
            self._sheet = self._wb.create_sheet()
        self._sheet.title = name

    def _get_cell(self, value, cell_format):
        cell = WriteOnlyCell(self._sheet, value)
        cell.number_format = cell_format
        return cell

    async def write_conc(self, amodel: ConcActionModel, data: KwicPageData, args: SaveConcArgs):
        self._sheet.title = amodel.plugin_ctx.translate('concordance')
        self._import_row = lang_row_to_list

        # Determine keys first
        if len(data.Lines) > 0:
            if 'Left' in data.Lines[0]:
                left_key, kwic_key, right_key = 'Left', 'Kwic', 'Right'
            elif 'Sen_Left' in data.Lines[0]:
                left_key, kwic_key, right_key = 'Sen_Left', 'Kwic', 'Sen_Right'
            else:
                raise ConcordanceQueryParamsError(amodel.translate('Invalid data'))
        else:
            left_key, kwic_key, right_key = 'Left', 'Kwic', 'Right'

        # Validate split character
        if args.split_content_by_char and not args.split_char:
            raise ConcordanceQueryParamsError(amodel.translate('Split character cannot be empty'))

        # Calculate number of split columns if splitting is enabled
        num_split_cols = 0
        if args.split_content_by_char and args.split_char and len(data.Lines) > 0:
            # Calculate max split columns across all lines to ensure consistent column count
            num_split_cols = self._calculate_max_split_columns(
                data.Lines,
                args.split_char,
                lambda line: self._merge_conc_line_parts(
                    line[kwic_key], data.merged_attrs, amodel.args.attr_vmode not in ['mouseover']
                )
            )

        if args.heading:
            doc_struct = amodel.corp.get_conf('DOCSTRUCTURE')
            refs_args = [x.strip('=') for x in amodel.args.refs.split(',')]
            used_refs = [
                ('#', amodel.plugin_ctx.translate('Token number')),
                (doc_struct, amodel.plugin_ctx.translate('Document number')),
                *[(x, x) for x in amodel.corp.get_structattrs()],
            ]
            used_refs = [x[1] for x in used_refs if x[0] in refs_args]
            # Add extra columns for split parts
            base_headers = [''] + used_refs if args.numbering else used_refs
            # Structure: [numbering?] + refs + left_context + kwic + [split_cols] + right_context
            # We need to insert split columns after kwic, which comes after refs and left_context
            if num_split_cols > 0:
                # Build headers: base_headers (numbering + refs) + left + kwic + splits + right
                ref_headers = base_headers + ['', ''] + [''] * num_split_cols + ['']
            else:
                ref_headers = base_headers + ['', '', '']
            self._write_ref_headings(ref_headers)

        for row_num, line in enumerate(data.Lines, args.from_line):
            lang_rows = self._process_lang(
                line, left_key, kwic_key, right_key, amodel.lines_groups.is_defined(), amodel.args.attr_vmode, data.merged_attrs, data.merged_ctxattrs)
            if 'Align' in line:
                lang_rows += self._process_lang(
                    line['Align'], left_key, kwic_key, right_key, False, amodel.args.attr_vmode, data.merged_attrs, data.merged_ctxattrs)
            
            # Apply split if enabled and pad to max columns
            if args.split_content_by_char and args.split_char:
                lang_rows = self._split_kwic_if_needed(lang_rows, args.split_char)
                # Pad rows to ensure consistent column count
                for lang_row in lang_rows:
                    for i in range(1, num_split_cols + 1):
                        if f'kwic_split_{i}' not in lang_row:
                            lang_row[f'kwic_split_{i}'] = ''
            
            self._writerow(row_num + args.numbering_offset if args.numbering else None, *lang_rows)

    async def write_coll(self, amodel: ConcActionModel, data: CalculateCollsResult, args: SavecollArgs):
        self._sheet.title = amodel.plugin_ctx.translate('collocations')
        
        # Validate split character
        if args.split_content_by_char and not args.split_char:
            raise ConcordanceQueryParamsError(amodel.translate('Split character cannot be empty'))

        # Calculate number of split columns if splitting is enabled
        num_split_cols = 0
        if args.split_content_by_char and args.split_char and len(data.Items) > 0:
            # Calculate max split columns across all items to ensure consistent column count
            num_split_cols = self._calculate_max_split_columns(
                data.Items,
                args.split_char,
                lambda item: item['str']
            )
        
        # Adjust column types for split columns
        if num_split_cols > 0:
            self._set_col_types(int, *([str] * (num_split_cols + 1)), *((float,) * 8))
        else:
            self._set_col_types(int, str, *((float,) * 8))
        
        if args.colheaders or args.heading:
            # Adjust headers for split columns
            if num_split_cols > 0:
                # First column is the collocation (split into multiple), then freq and stats
                headers = [''] + [''] * num_split_cols + [item['n'] for item in data.Head]
            else:
                headers = [''] + [item['n'] for item in data.Head]
            self._writeheading(headers)
        
        for i, item in enumerate(data.Items, 1):
            # Apply split if enabled
            if args.split_content_by_char and args.split_char:
                coll_str = item['str']
                split_parts = coll_str.split(args.split_char)
                # Pad to max columns to ensure consistent column count
                while len(split_parts) - 1 < num_split_cols:
                    split_parts.append('')
                row_data = tuple(split_parts) + (item['freq'],) + tuple(stat['s'] for stat in item['Stats'])
            else:
                row_data = (item['str'], item['freq'], *(stat['s'] for stat in item['Stats']))
            
            self._writerow(i, row_data)

    async def write_freq(self, amodel: ConcActionModel, data: Dict[str, Any], args: SavefreqArgs):
        self._sheet.title = amodel.plugin_ctx.translate('frequency distribution')

        # Here we expect that when saving multi-block items, all the block have
        # the same number of columns which is quite bad. But currently there is
        # no better common 'denominator'.
        num_word_cols = len(data['Blocks'][0].get('Items', [{'Word': []}])[0].get('Word'))
        self._set_col_types(*([int] + num_word_cols * [str] + [float, float]))

        for block in data['Blocks']:
            if args.multi_sheet_file:
                self._new_sheet(block['Head'][0]['n'])

            if args.colheaders or args.heading:
                self._writeheading([''] + [item['n'] for item in block['Head'][:-2]] +
                                   ['freq', 'freq [%]'])
            for i, item in enumerate(block['Items'], 1):
                self._writerow(i, [w['n'] for w in item['Word']] +
                               [item['freq'], item.get('rel', '')])

    async def write_pquery(self, amodel: ParadigmaticQueryActionModel, data: List[PqueryDataLine], args: SavePQueryArgs):
        self._sheet.title = amodel.plugin_ctx.translate('Paradigmatic query')
        freq_cols = len(data[0].freqs)
        self._set_col_types(int, str, *(float for _ in range(freq_cols)))
        if args.colheaders or args.heading:
            self._writeheading(['', 'value', *(f'freq{i+1}' for i in range(freq_cols)), 'freq'])

        for i, row in enumerate(data, 1):
            self._writerow(i, (row.value, *row.freqs, sum(row.freqs)))

    async def write_keywords(self, amodel: KeywordsActionModel, result: KeywordsResult, args: SaveKeywordsArgs):
        self._sheet.title = amodel.plugin_ctx.translate('Keyword analysis')
        if args.colheaders or args.heading:
            if isinstance(result.data[0], CNCKeywordLine):
                self._set_col_types(int, str, float, float, float, int, int, float, float)
                self._writeheading(['', 'item', 'logL', 'chi2', 'din',
                                    'frq', 'frq_ref', 'ipm', 'ipm_ref',])
            else:
                self._set_col_types(int, str, float, int, int, float, float)
                self._writeheading(['', 'item', 'score', 'freq', 'frq_ref', 'ipm', 'ipm_ref'])

        for i, row in enumerate(result.data, 1):
            if isinstance(row, CNCKeywordLine):
                self._writerow(i, (row.item, row.logL, row.chi2, row.din,
                                   row.frq1, row.frq2, row.rel_frq1, row.rel_frq2))
            else:
                self._writerow(i, (row.item, row.score, row.frq1,
                                   row.frq2, row.rel_frq1, row.rel_frq2))

    async def write_wordlist(self, amodel: WordlistActionModel, data: List[Tuple[str, int]], args: WordlistSaveFormArgs):
        self._sheet.title = amodel.plugin_ctx.translate('word list')
        self._set_col_types(int, str, float)

        if args.colheaders:
            self._writeheading(['', amodel.curr_wlform_args.wlattr, 'freq'])

        elif args.heading:
            self._writeheading([
                'corpus: {}\nsubcorpus: {}\npattern: {}'.format(
                    amodel.corp.human_readable_corpname, amodel.args.usesubcorp, amodel.curr_wlform_args.wlpat),
                '', ''
            ])

        for i, (wlattr, freq) in enumerate(data, 1):
            self._writerow(i, (wlattr, str(freq)))


def create_instance(locale: Locale):
    return XLSXExport(locale)
